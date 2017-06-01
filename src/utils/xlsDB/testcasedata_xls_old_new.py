# -*- coding: utf-8 -*-
import os,re,string
import xlrd
# from robot.errors import DataError, VariableError
import json
#from xlutils.copy import copy
import xlwt

from openpyxl import Workbook,load_workbook
from collections import OrderedDict

# from django.collection import OrderedDict



def GetFileList(tag,path):
    """
    @name    :  GetFileList
    @brief   :  Get file<tag> list
    @parem   :  <tag><path>
                <tag> : File extension
                <path>: File Directory
    @return  :  FileList
    @pkgpath :  
    @remark  :  
    @author  :  Jason
    @history :  Create 2016/5/2
    """
    FileList = []

    if re.match(".*\$RECYCLE\.BIN$",path) != None or re.match(".*System Volume Information$",path) != None:
        return FileList

    #print "path is %s,  os.path.isdir(path) is %s " %(path,os.path.isdir(path))
    if os.path.isdir(path) == True :
        if path[-1:] != '/' and path[-1:] != '\\' :
            path = path + '/'
        PathList = os.listdir(path)
        #print "PathList1 is  %s"  %(PathList)
        for file in PathList:
            FileList = FileList + GetFileList(tag,path+file)
            #print  "FileList2 is %s"  %(FileList)
    else :
        if os.path.isfile(path) == True :
            if os.path.splitext(path)[1][1:] == tag and  ( '~' not in os.path.splitext(path)[0]):
                FileList.append(path)
    return FileList




def xls_old_new(old_name,play_name):
    returnData={}
    #path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) + "/testcase/CP_PaiCai/"
#    filename11='E:/work/svn/01CP/11AutoTest/00RobotFramework/ISRProject/testcase/CP_PaiCai/CP_PC_'+old_name+'.xlsx'
#    filename99='E:/work/svn/01CP/11AutoTest/00RobotFramework/ISRProject/testcase/CP_PaiCai/CP_PC_'+play_name+'.xlsx'
#     filename11 = path +'/' + old_name
#     #print filename11
#     filename99 = path +'/'+ play_name
    filename11 = old_name
    #print filename11
    filename99 = play_name

    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = load_workbook(filename11)

    sheetnames = wb.get_sheet_names()
    #if play_name in filename99:
    for i in range(len(sheetnames)):
#     for i in range(1):
        print 'iiii___',i
        ws = wb.get_sheet_by_name(sheetnames[i])
        nrows=ws.max_row
        ncols = ws.max_column
#         rows=table.max_row   #获取行数
# 
#         cols=table.max_column    #获取列数
#         title_list = ws.title.split('_')
#         if title_list[0]==play_name:
#             pass
#         elif play_name in title_list:
#             print title_list
#         else:
#             title_list[0]=play_name
#         ws.title = '_'.join(title_list)
#         colname = ws.cell(row=0, column=j).value
#         rowstag=ws.row(0).value
        rowstag = []
        for r in range(1,ncols+1):
            rowstag.append(ws.cell(row=1,column=r).value)
#         rowstag.append(ws.columns)
#         print 'rowstag',rowstag
                                 
        for nc in range(1,ncols+1):
#             ws.cell(row=j + 1, column=1).value = 'CP_PC_'+ ws.title+'_'+str(j)
            #print 'play_name----',play_name,dict_list[play_name]
            colname = ws.cell(row=1, column=nc).value
         # model    model1    model2    caseID    caseName    normal    caseLevel    preRequisite    caseStep    expresult    caseData    remarks    testEnvironment    autoSupport    author    status    testResult
  
            if colname == u'模块':
                ws.cell(row=1, column=nc).value = 'model'
            elif colname== u'一级模块':
                ws.cell(row=1, column=nc).value = 'model1'
            elif colname == u'二级模块':
                ws.cell(row=1, column=nc).value='model2'
            elif colname == u'用例编号':
                ws.cell(row=1, column=nc).value = 'caseID'
            elif colname == u'用例名称':
                ws.cell(row=1, column=nc).value = 'caseName'
            elif colname == u'用例级别':
                ws.cell(row=1, column=nc).value = 'caseLevel'
            elif colname == u'预制条件':
                ws.cell(row=1, column=nc).value = 'preRequisite'
            elif colname == u'测试步骤':
                ws.cell(row=1, column=nc).value = 'caseStep'
            elif colname == u'预期结果':
                ws.cell(row=1, column=nc).value = 'expresult'
            elif colname == u'测试数据' or colname =='caseData':
                ws.cell(row=1, column=nc).value = 'caseData'
                CaseVar = nc
                print 'naum-->',CaseVar
            elif colname == u'备注':
                ws.cell(row=1, column=nc).value = 'remarks'
            elif colname == u'测试环境':
                ws.cell(row=1, column=nc).value = 'testEnvironment'
            elif colname == u'自动化支持' or colname=='autoSupport':
                ws.cell(row=1, column=nc).value = 'autoSupport'
                CaseVar_auto=nc
            elif colname == u'作者':
                ws.cell(row=1, column=nc).value = 'author'
            elif colname == u'测试结果':
#                 print nc
                ws.cell(row=1, column=nc).value = 'testResult'
                
            
#             if 'caseData' == colname :
#             if u'测试数据' == colname :
#                 CaseVar = nc
#                 print 'CaseVar',CaseVar
#         for j in range(1, nrows):

#         for i in range(1, 2):
            
#         for j in range(1, nrows):
# #             returnData[i]=json.dumps(dict(zip(rowstag,ws.cell(row=i, column=j).value)))
# #                 #通过编解码还原数据
# #             returnData[i]=json.loads(returnData[i])
# #             
# #             print 'returnData999___>>>>',returnData
# #             print 'j',j ,ws.cell(row=j + 1, column=CaseVar).value
#             print j,ws.cell(row=j+1, column=4).value
#                 
#             Varstrlist = ws.cell(row=j + 1, column=CaseVar).value.split('\n')
# #             print Varstrlist
# #             VarDict = {}
#             VarDict = OrderedDict()
#             #map(lambda x: VarDict.setdefault(x.split('=')[0],x.split('=')[1]), Varstrlist)
#             for Varstr in Varstrlist:
#                 h=0
#                 if '' != Varstr :
#                     try:
#                         l_index = Varstr.find('=')
# #                         VarDict.setdefault(Varstr[0:l_index],Varstr[l_index+1:])
# #                         s1 = Varstr[l_index+1:]
# #                         s2 = [unicode(s).encode('utf-8') for s in s1]  # utf-8 encoding
# #                             VarDict.setdefault(Varstr[0:l_index],unicode(Varstr[l_index+1:]).encode('utf-8'))
#                         VarDict[Varstr[0:l_index]]=Varstr[l_index+1:]
# #                         VarDict.setdefault(Varstr[0:l_index], Varstr[l_index+1:])
# #                             print 'ikkkk'
# #                             VarDict[h]=json.dumps(dict(zip([Varstr[0:l_index]],[Varstr[l_index+1:]])))
# #                             #通过编解码还原数据
# #                             print 'noo'
# #                             VarDict[h]=json.loads(VarDict[h])
# #                             print '3333'
#                         #VarDict.setdefault(Varstr.split('=')[0],Varstr.split('=')[1])
#                     except:
#                         raise ("TEST-DATA ERROR: %s" % str(Varstrlist))
#                 h += 1
# #             print  VarDict
# #             print str(VarDict).decode('raw_unicode_escape')
#         
# #             s1 = sheet.row_values(col)
# #             s2 = [unicode(s).encode('utf-8') for s in s1]  # utf-8 encoding
# #             data.append(dict(zip(title, s2)))
# 
#             ws.cell(row=j + 1, column=CaseVar).value = str(VarDict).replace("OrderedDict([(u'",'{"').replace("')])",'"}').replace("', u'",'": "').replace("'), (u'",'", "').decode('raw_unicode_escape')  # 重新写入测试数据
# #             ws.cell(row=j + 1, column=CaseVar).value = str(VarDict).replace("{u'",'{"').replace("'}",'"}').replace("': u'",'": "').replace("', u'",'", "').decode('raw_unicode_escape')  # 重新写入测试数据
# #             numn_index=dict_list['FC3D'][1:].index(int(ws.cell(row=j + 1, column=4).value))   #找出 methodid在list[1:]中的索引值
# #             ws.cell(row=j + 1, column=4).value=dict_list[play_name][1:][numn_index]          #替换methodid在list[1:]中的值
#             
#             if ws.cell(row=j+1, column=CaseVar_auto).value ==u'是':
#                 ws.cell(row=j+1, column=CaseVar_auto).value='Y'
#             elif ws.cell(row=j+1, column=CaseVar_auto).value ==u'否':
#                 ws.cell(row=j+1, column=CaseVar_auto).value=''

    wb.save(filename99)   # 最后一定要保存！
    


if __name__ == "__main__":
#     path = "D:\workspace\web_autotest_my\src\data"
    path = r"C:\work\svn\Doc\01CP\11AutoTest\web_autotest_project\src\data\web_old"
    xls_list=GetFileList('xlsx',path) + GetFileList('xls',path)
    
    for xls_i in xls_list:
        print xls_i
        xls_old_new(xls_i, xls_i)
#     print 'xls_list--->',xls_list
#     xls_old_new("CP_WEB.ZCDL_JASON.xlsx", "CP_WEB.ZCDL_JASON_new.xlsx")


    
